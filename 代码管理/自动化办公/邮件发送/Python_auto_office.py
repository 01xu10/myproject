# !/usr/bin/env python
# -*- coding: utf-8 -*-


import smtplib, time
from openpyxl import load_workbook
from email.mime.text import MIMEText
from email.header import Header

def main():
    # 1、登录STMP服务器，需要使用QQ邮箱动态授权码！
    smtp_obj = smtplib.SMTP_SSL('smtp.qq.com', 465)                            # 发件人的SMTP邮件服务器，端口是465
    smtp_obj.login('@qq.com', '')                     # 获取动态授权码
    # smtp_obj.set_debuglevel(1)                                               # 显示调试信息

    # 2、打开工资条excel文件，并得到活跃sheet表
    wb = load_workbook('员工工资表.xlsx', data_only=True)
    sheet = wb['Sheet1']

    # 3、循环发送excel表格的行，按行发送
    count = 0
    table_head = '<thead>'
    for row in sheet.iter_rows(min_row=1, max_row=16):
        # 3.1 将第一行表头拼接
        count += 1
        if count == 1:
            for item in row:
                table_head += f'<th>{item.value}</th>'
            table_head += "</thead>"
            continue

        else:
            # 3.2 将每一行工资内容，以表格形式拼接
            row_content = '<tr>'
            for cell in row:
                row_content += f'<td>{cell.value}</td>'
            row_content += '</tr>'

            # 3.3 获取员工的Email
            staff_email = row[1].value

            # 3.4 获取员工的姓名
            staff_name = row[2].value

        # 4、构造要发送的邮件
        # 4.1 邮件的内容
        msg_content = \
           f'''
                <div>
                    <h2>{staff_name}</h2>
                    <p style="font-size: 20px">&emsp;&emsp;你好：请查收你2020年12月的工资条, 内容如下：</p>
                    <table border="1px solid black" style="font-size: 15px">
                        {table_head}
                        {row_content}
                    </table>
                </div>
            '''
        # 4.2 邮件正文，使用html格式发送
        message = MIMEText(msg_content, 'html', 'utf-8')
        # 4.3 邮件头：发送者信息
        message['From'] = Header('Vce', 'utf-8')
        # 4.4 邮件头：接收者信息
        message['To'] = Header(staff_name, 'utf-8')
        # 4.5 邮件头：邮件主题信息
        message['Subject'] = Header('贝塔哥哥发工资啦', 'utf-8')

        # 5、发送邮件，并打印发送的结果！
        try:
            time.sleep(5)
            smtp_obj.sendmail('@qq.com', [staff_email], message.as_string())  # 压缩成字符串格式发送
            print('***成功发送工资条：{}***'.format(staff_name))
        except Exception as e:
            print('***邮件发送错误***')


if __name__ == '__main__':
    main()
