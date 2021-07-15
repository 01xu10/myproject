# !/usr/bin/env python
# -*- coding: utf-8 -*-
# ------------------------------

import smtplib,time
from openpyxl import load_workbook
from email.mime.text import MIMEText
from email.header import Header

class E_mail(object):
    # 1、登录STMP服务器，需要使用QQ邮箱动态授权码！
    def __init__(self):
        # 发件人的SMTP邮件服务器，端口是465
        self.smtp_obj = smtplib.SMTP_SSL('smtp.qq.com', 465)
        # 获取动态授权码
        self.smtp_obj.login('1150772265@qq.com', 'xmzgurvlhyekgcag')
        # 显示调试信息
        # self.smtp_obj.set_debuglevel(1)
        # 计数器
        self.count = 0
        # 工资头部
        self.table_head = '<thead>'

    # 2.打开工资条excel文件，并得到活跃表sheet表
    def open_excel(self):
        # data_only只显示数据，隐藏函数
        wb = load_workbook('员工工资表.xlsx', data_only=True)
        sheet = wb['Sheet1']
        for row in sheet.iter_rows(min_row=1, max_row=16):
            self.operate_excel(row)


    # 3.循环发送excel表格的行，按行发送
    def operate_excel(self, row):
        # 3.1将表头取出进行拼接
        self.count += 1
        if self.count == 1:
            self.table_head = '<thead>'
            for item in row:
                self.table_head += f'<th>{item.value}</th>'
            self.table_head += '</thead>'

        # 3.2 将每一行工资内容，以表格形式拼接
        else:
            row_content = '<tr>'
            for cell in row:
                row_content += f'<td>{cell.value}</td>'
            row_content += '</tr>'

            # 3.3 获取员工的Email
            staff_email = row[1].value

            # 3.4 获取员工的姓名
            staff_name = row[2].value

            self.msg_content(self.table_head, row_content, staff_email, staff_name)


    # 4.构造要发送邮件的内容
    def msg_content(self,table_head, row_content, staff_email, staff_name):
        # 4.1 邮件的内容
        msg_content = \
            f'''
                <div>
                    <h2>{staff_name}</h2>
                    <p style="font-size: 20px">&emsp;&emsp;你好：请查收你2021年1月的工资条, 内容如下：</p>
                    <table border="1px solid black" style="font-size: 15px">
                        {table_head}
                        {row_content}
                    </table>
                </div>
            '''
        # 4.2 邮件正文，使用html格式发送
        message = MIMEText(msg_content, 'html', 'utf-8')

        # 4.3 邮件头：发送者信息
        message['From'] = Header('HR', 'utf-8')

        # 4.4 邮件头：接收者信息
        message['To'] = Header(staff_name, 'utf-8')

        # 4.5 邮件头：邮件主题信息
        message['Subject'] = Header('这个月的工资到账啦！', 'utf-8')

        self.send_email(msg_content, message, staff_name, staff_email)


    # 5.发送邮件
    def send_email(self, msg_content, message, staff_name, staff_email):
        try:
            time.sleep(5)
            # 压缩成字符串格式发送
            self.smtp_obj.sendmail('1150772265@qq.com',staff_email,message.as_string())
            print('***成功发送工资条：{}***'.format(staff_name))
        except Exception as e:
            print('***邮件发送错误,员工：{}发送失败***'.format(staff_name))



    # 6.执行程序
    def main(self):
        self.open_excel()


if __name__ == '__main__':
    email = E_mail()
    email.main()

