# !/usr/bin/env python
# -*- coding: utf-8 -*-
# ------------------------------
# @project : Py_Project
# @File    : HUPU_NBA_Spider.py
# @Software: PyCharm
# @Author  : Xu
# @Time    : 2021/2/12 19:31
# ------------------------------

import requests
from lxml import etree
from xlutils.copy import copy
import os, xlwt, xlrd
from pprint import pprint
from openpyxl import Workbook


class NBA_numbers(object):
    # 1.url+headers+excel
    def __init__(self):
        self.start_url = r'https://nba.hupu.com/players'
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_3) AppleWebKit/537.36 (KHTML, like Gecko) '
                          'Chrome/87.0.4280.88 Safari/537.36'
        }
        # self.wb = Workbook()
        # self.wb.remove(self.wb['Sheet'])
        self.num = 0
        self.worksheet1 = {}


    # 2.发起请求
    def request(self):
        respone = requests.get(self.start_url, headers=self.headers).text
        html_str = etree.HTML(respone)
        nba_urls = html_str.xpath(r'//ul/li/span/a/@href')
        nba_teams = html_str.xpath(r'//ul/li/span/a/text()')
        # 3.对每个球队发起请求
        for nba_url,nba_team in zip(nba_urls,nba_teams):
            # print(nba_url, nba_team, self.num)
            # 创建每个球队的sheet
            # self.wb.create_sheet(title=nba_team)
            # self.sheet = self.wb[nba_team]
            # self.sheet.append(['姓名', '英文名', '号码', '位置', '身高', '体重', '生日', '合同'])
            self.analysis_data(nba_url, nba_team)
            # self.num += 1

    # 3.分析数据，爬取NBA球员的个人信息
    def analysis_data(self,nba_url,nba_team):
        respone = requests.get(nba_url,headers=self.headers).text
        html_str = etree.HTML(respone)
        nba_names= html_str.xpath(r'//td[2]/b/a/text()')
        nba_Eng_names = html_str.xpath(r'//td[2]/p/b/text()')
        nba_nums = html_str.xpath(r'//td[3]/text()')[1::]
        nba_adds = html_str.xpath(r'//td[4]/text()')[1::]
        nba_heights = html_str.xpath(r'//td[5]/text()')[1::]
        nba_weights = html_str.xpath(r'//td[6]/text()')[1::]
        nba_birs = html_str.xpath(r'//td[7]/text()')[1::]
        nba_coms = html_str.xpath(r'//td[8]/text()')[1::]
        nba_ms = html_str.xpath(r'//td[8]/b/text()')
        for nba_name,nba_Eng_name,nba_num,nba_add,nba_height,nba_weight,nba_bir,nba_com,nba_m in zip(nba_names,nba_Eng_names,nba_nums,nba_adds,nba_heights, nba_weights, nba_birs, nba_coms, nba_ms):
            #1111
            self.num += 1
            nba_play_list = [nba_name, nba_Eng_name,nba_num,nba_add,nba_height,nba_weight,nba_bir,nba_com,nba_m]
            data = {
                nba_team: nba_play_list
            }
            self.save_excel(data, nba_team)
            print('\r***已下载完成了：{}个球员信息***'.format(self.num), end='')

            # self.save_imfo(nba_name, nba_Eng_name,nba_num,nba_add,nba_height,nba_weight,nba_bir,nba_com,nba_m)

    def save_excel(self, data, nba_team):
        # data = {
        #     '基本详情': ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j']
        # }
        os_path_1 = os.getcwd() + '/NBA球员信息/'
        if not os.path.exists(os_path_1):
            os.mkdir(os_path_1)
        # os_path = os_path_1 + self.os_path_name + '.xls'
        os_path = os_path_1 + 'HUPU_NBA.xls'
        if not os.path.exists(os_path):
            # 创建新的workbook（其实就是创建新的excel）
            workbook = xlwt.Workbook(encoding='utf-8')
            # 创建新的sheet表
            self.worksheet1[nba_team] = workbook.add_sheet(nba_team, cell_overwrite_ok=True)
            borders = xlwt.Borders()  # Create Borders
            """定义边框实线"""
            borders.left = xlwt.Borders.THIN
            borders.right = xlwt.Borders.THIN
            borders.top = xlwt.Borders.THIN
            borders.bottom = xlwt.Borders.THIN
            borders.left_colour = 0x40
            borders.right_colour = 0x40
            borders.top_colour = 0x40
            borders.bottom_colour = 0x40
            style = xlwt.XFStyle()  # Create Style
            style.borders = borders  # Add Borders to Style
            """居中写入设置"""
            al = xlwt.Alignment()
            al.horz = 0x02  # 水平居中
            al.vert = 0x01  # 垂直居中
            style.alignment = al
            # 合并 第0行到第0列 的 第0列到第13列
            '''基本详情13'''
            # worksheet1.write_merge(0, 0, 0, 13, '基本详情', style)
            excel_data_1 = ('姓名', '英文名', '号码', '位置', '身高', '体重', '生日', '合同')
            for i in range(0, len(excel_data_1)):
                self.worksheet1[nba_team].col(i).width = 2560 * 3
                #               行，列，  内容，            样式
                self.worksheet1[nba_team].write(0, i, excel_data_1[i], style)
            workbook.save(os_path)
        # 判断工作表是否存在
        if os.path.exists(os_path):
            # 打开工作薄
            workbook = xlrd.open_workbook(os_path)
            # 获取工作薄中所有表的个数
            sheets = workbook.sheet_names()
            for i in range(len(sheets)):
                for name in data.keys():
                    worksheet = workbook.sheet_by_name(sheets[i])
                    # 获取工作薄中所有表中的表名与数据名对比
                    if worksheet.name == name:
                        # 获取表中已存在的行数
                        rows_old = worksheet.nrows
                        # 将xlrd对象拷贝转化为xlwt对象
                        new_workbook = copy(workbook)
                        # 获取转化后的工作薄中的第i张表
                        new_worksheet = new_workbook.get_sheet(i)
                        for num in range(0, len(data[name])):
                            new_worksheet.write(rows_old, num, data[name][num])
                        new_workbook.save(os_path)

    # 4.保存数据
    # def save_imfo(self,nba_name, nba_Eng_name,nba_num,nba_add,nba_height,nba_weight,nba_bir,nba_com,nba_m):
    #     self.sheet.append([nba_name, nba_Eng_name,nba_num,nba_add,nba_height,nba_weight,nba_bir,nba_com+nba_m])

    # 5.执行函数
    def main(self):
        self.request()
        # self.wb.save('NBA_player_info.xlsx')

if __name__ == '__main__':
    numbers = NBA_numbers()
    numbers.main()
